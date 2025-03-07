import re
import csv
import logging
from bs4 import BeautifulSoup
from typing import List

logger = logging.getLogger("xml_processor")

def process_section(section):
    md = ""
    head = section.find('head')
    if head and head.text:
        md += f"## {head.text.strip()}\n\n"
    for block in section.find_all('block', recursive=False):
        md += process_block(block)
    return md

def process_block(block):
    md = ""
    md += process_block_head(block)
    md += process_block_notices(block)
    md += process_block_figures(block)
    md += process_block_paragraphs(block)
    md += process_block_lists(block)
    md += process_block_tables(block)
    md += process_block_subblocks(block)
    md += "\n"
    return md

def process_block_head(block):
    md = ""
    head = block.find('head')
    if head and head.text:
        md += f"### {head.text.strip()}\n\n"
    return md

def process_block_notices(block):
    md = ""
    for notice_type in ['warning', 'caution', 'important', 'note']:
        for notice in block.find_all(notice_type, recursive=False):
            md += process_notice(notice, notice_type)
    return md

def process_block_figures(block):
    md = ""
    for figure in block.find_all('figure', recursive=False):
        md += process_figure(figure)
    return md

def process_block_paragraphs(block):
    md = ""
    for para in block.find_all('para', recursive=False):
        md += process_paragraph(para)
    return md

def process_block_lists(block):
    md = ""
    for list_elem in block.find_all('list', recursive=False):
        md += process_list(list_elem)
    for proclist in block.find_all('proclist', recursive=False):
        md += process_proclist(proclist)
    return md

def process_block_tables(block):
    md = ""
    for table in block.find_all('table', recursive=False):
        md += process_table(table)
    return md

def process_block_subblocks(block):
    md = ""
    for subblock in block.find_all('subblock', recursive=False):
        subhead = subblock.find('head')
        if subhead and subhead.text:
            md += f"#### {subhead.text.strip()}\n\n"
        for child in subblock.children:
            if not hasattr(child, 'name'):
                continue
            if child.name == 'para':
                md += process_paragraph(child)
            elif child.name == 'list':
                md += process_list(child)
            elif child.name == 'proclist':
                md += process_proclist(child)
            elif child.name == 'figure':
                md += process_figure(child)
            elif child.name in ['warning', 'caution', 'important', 'note']:
                md += process_notice(child, child.name)
    return md

def process_paragraph(para):
    text = extract_text(para)
    if not text.strip():
        return ""
    return f"{text}\n\n"

def process_list(list_elem):
    md = ""
    list_type = list_elem.get('type', 'unordered')
    for item in list_elem.find_all('item'):
        item_para = item.find('para')
        if item_para:
            item_text = extract_text(item_para)
            if list_type == 'unordered':
                md += f"- {item_text}\n"
            else:
                md += f"1. {item_text}\n"
    md += "\n"
    return md

def process_proclist(proclist):
    md = ""
    for i, step in enumerate(proclist.find_all('step'), 1):
        step_text = ""
        for notice_type in ['warning', 'caution', 'important', 'note']:
            for notice in step.find_all(notice_type):
                step_text += process_notice(notice, notice_type)
        for para in step.find_all('para', recursive=False):
            para_text = extract_text(para)
            if para_text:
                step_text += para_text + "\n\n"
        for list_elem in step.find_all('list'):
            step_text += process_list(list_elem)
        md += f"{i}. {step_text}\n"
    md += "\n"
    return md

def process_figure(figure):
    md = ""
    caption = figure.find('caption')
    caption_text = ""
    if caption and caption.get('print') == '0' and caption.text:
        caption_text = caption.text.strip()
    for graphic in figure.find_all('graphic'):
        alt_text = caption_text if caption_text else "Image"
        md += f"![{alt_text}]\n"
    if caption_text:
        md += f"*{caption_text}*\n"
    legend = figure.find('legend')
    if legend:
        md += "*Legend:*\n"
        terms = legend.find_all('term')
        defs = legend.find_all('def')
        for i, term in enumerate(terms):
            if i < len(defs):
                term_text = term.text.strip()
                def_text = defs[i].text.strip()
                md += f"- {term_text}: {def_text}\n"
    md += "\n"
    return md

def process_notice(notice, notice_type):
    md = ""
    nested = notice.find(['caution', 'important', 'note'])
    if nested:
        notice_type = nested.name.upper()
        notice = nested

    notice_text = ""
    for para in notice.find_all('para'):
        para_text = extract_text(para)
        if para_text:
            notice_text += para_text + "\n\n"

    if notice_text:
        md += f"**{notice_type.upper()}**: {notice_text}\n"
    return md

def process_table(table):
    md = ""
    caption = table.find('caption')
    if caption and caption.get('print') == '0' and caption.text:
        md += f"*{caption.text.strip()}*\n\n"

    tgroup = table.find('tgroup')
    if not tgroup:
        logger.warning("Table has no tgroup element.")
        return md

    colspecs = tgroup.find_all('colspec')
    col_names = {}
    for i, colspec in enumerate(colspecs):
        colname = colspec.get('colname')
        if colname:
            col_names[colname] = i

    total_cols = len(colspecs) if colspecs else 0
    if total_cols == 0:
        first_row = tgroup.find('row')
        total_cols = len(first_row.find_all('entry')) if first_row else 2

    header_processed = False

    thead = tgroup.find('thead')
    if thead:
        rows = thead.find_all('row')
        for row in rows:
            header = "|"
            entries = row.find_all('entry')
            processed_cols = set()
            for entry in entries:
                start_col = entry.get('namest')
                end_col = entry.get('nameend')
                span = 1
                if start_col and end_col and start_col in col_names and end_col in col_names:
                    start_idx = col_names[start_col]
                    end_idx = col_names[end_col]
                    span = end_idx - start_idx + 1
                    for i in range(start_idx, end_idx + 1):
                        processed_cols.add(i)
                else:
                    for i in range(total_cols):
                        if i not in processed_cols:
                            processed_cols.add(i)
                            break
                cell_text = extract_text(entry).replace("\n", " ")
                if span > 1:
                    header += f" {cell_text} |" + " |" * (span - 1)
                else:
                    header += f" {cell_text} |"
            md += f"{header}\n"
            if row == rows[0]:
                separator = "|"
                for _ in range(total_cols):
                    separator += "------------|"
                md += f"{separator}\n"
        header_processed = True

    tbody = tgroup.find('tbody')
    if tbody:
        rows = tbody.find_all('row')
        if not rows:
            logger.warning("Table has no rows.")
            return md
        if not header_processed and len(rows) > 0:
            header_row = rows[0]
            header = "|"
            entries = header_row.find_all('entry')
            processed_cols = set()
            for entry in entries:
                start_col = entry.get('namest')
                end_col = entry.get('nameend')
                span = 1
                if start_col and end_col and start_col in col_names and end_col in col_names:
                    start_idx = col_names[start_col]
                    end_idx = col_names[end_col]
                    span = end_idx - start_idx + 1
                    for i in range(start_idx, end_idx + 1):
                        processed_cols.add(i)
                else:
                    for i in range(total_cols):
                        if i not in processed_cols:
                            processed_cols.add(i)
                            break
                cell_text = extract_text(entry).replace("\n", " ")
                if span > 1:
                    header += f" {cell_text} |" + " |" * (span - 1)
                else:
                    header += f" {cell_text} |"
            md += f"{header}\n"
            separator = "|"
            for _ in range(total_cols):
                separator += "------------|"
            md += f"{separator}\n"
            rows = rows[1:]
        for row in rows:
            row_text = "|"
            entries = row.find_all('entry')
            processed_cols = set()
            for entry in entries:
                start_col = entry.get('namest')
                end_col = entry.get('nameend')
                span = 1
                if start_col and end_col and start_col in col_names and end_col in col_names:
                    start_idx = col_names[start_col]
                    end_idx = col_names[end_col]
                    span = end_idx - start_idx + 1
                    for i in range(start_idx, end_idx + 1):
                        processed_cols.add(i)
                else:
                    for i in range(total_cols):
                        if i not in processed_cols:
                            processed_cols.add(i)
                            break
                cell_text = extract_text(entry).replace("\n", " ")
                if span > 1:
                    row_text += f" {cell_text} |" + " |" * (span - 1)
                else:
                    row_text += f" {cell_text} |"
            md += f"{row_text}\n"
    md += "\n"
    return md

def extract_text(element):
    if not element:
        return ""
    result = ""
    for child in element.contents:
        if isinstance(child, str):
            result += child
        elif hasattr(child, 'name'):
            if child.name == 'emph' and child.get('etype') == 'bold':
                result += f"**{extract_text(child)}**"
            elif child.name == 'trademark':
                child_text = extract_text(child)
                tm_type = child.get('type', '')
                if tm_type == 'tm':
                    result += f"{child_text}™"
                elif tm_type == 'r':
                    result += f"{child_text}®"
                else:
                    result += child_text
            else:
                result += extract_text(child)
    return result

def markdown_to_csv(md_file_path, csv_file_path):
    
    def parse_md_row(line: str) -> List[str]:
        row = line.strip().strip('|')
        return [col.strip() for col in row.split('|')]

    with open(md_file_path, 'r', encoding='utf-8') as f:
        lines = f.readlines()

    tables = []
    current_table = []
    in_table = False

    for line in lines:
        if line.startswith("|"):
            in_table = True
            current_table.append(line.rstrip("\n"))
        else:
            if in_table and current_table:
                tables.append(current_table)
                current_table = []
            in_table = False

    if current_table:
        tables.append(current_table)

    with open(csv_file_path, 'w', newline='', encoding='utf-8') as csvfile:
        writer = csv.writer(csvfile)
        table_count = 1
        for table_lines in tables:
            if len(table_lines) < 1:
                continue
            # Write the table caption row as the first row.
            writer.writerow([f"{table_lines[0].strip().strip('*').strip()}"])
            # Next, check if second line is a separator.
            data_start_index = 1
            if len(table_lines) > 1 and re.match(r'^\|\s*-+', table_lines[1]):
                data_start_index = 2
            header_cols = parse_md_row(table_lines[data_start_index - 1])
            writer.writerow(header_cols)
            for row_line in table_lines[data_start_index:]:
                data_cols = parse_md_row(row_line)
                writer.writerow(data_cols)
            writer.writerow([])  # blank row between tables
            table_count += 1

def markdown_to_excel(md_file_path, excel_file_path):
    
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment
    except ImportError:
        logger.error("openpyxl is required for Excel output. Install it with 'pip install openpyxl'")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Tables"

    # Define styles for table names and cell wrapping.
    table_name_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    table_name_font = Font(color="FFFFFF", bold=True)
    wrap_alignment = Alignment(wrap_text=True)

    with open(md_file_path, 'r', encoding='utf-8') as f:
        lines = f.readlines()

    row_idx = 1
    i = 0
    while i < len(lines):
        line = lines[i].rstrip("\n")
        if line.startswith("*") and line.endswith("*"):
            caption = line.strip("*").strip()
            j = i + 1
            while j < len(lines) and not lines[j].strip():
                j += 1
            if j < len(lines) and lines[j].strip().startswith("|"):
                # Write the caption with formatting.
                cell = ws.cell(row=row_idx, column=1, value=caption)
                cell.fill = table_name_fill
                cell.font = table_name_font
                cell.alignment = wrap_alignment
                row_idx += 1
                i = j
                continue
            else:
                i += 1
                continue
        elif line.startswith("|"):
            # Start accumulating table rows.
            table_lines = []
            while i < len(lines) and lines[i].strip().startswith("|"):
                table_lines.append(lines[i].rstrip("\n"))
                i += 1
            if not table_lines:
                continue
            data_start = 1
            if len(table_lines) > 1 and re.match(r'^\|\s*-+', table_lines[1]):
                data_start = 2
            header = [col.strip() for col in table_lines[0].strip().strip("|").split("|")]
            col_count = len(header)
            for col_idx, col_val in enumerate(header, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=col_val)
                cell.alignment = wrap_alignment
            row_idx += 1
            for line in table_lines[data_start:]:
                cols = [col.strip() for col in line.strip().strip("|").split("|")]
                for col_idx, col_val in enumerate(cols, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=col_val)
                    cell.alignment = wrap_alignment
                row_idx += 1
            row_idx += 1
        else:
            i += 1

    wb.save(excel_file_path)
    logger.info(f"Excel conversion successful. Output saved to {excel_file_path}")