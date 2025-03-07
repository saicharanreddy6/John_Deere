import sys
import os
# Add the src folder to the Python module search path
sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "src"))

import unittest
import tempfile
import csv
import re
from openpyxl import load_workbook

# Import functions from your module.
# IMPORTANT: Ensure that the functions you want to test (e.g. markdown_to_csv) are defined at the module level.
from xml_process import (
    xml_to_markdown_conversion,
    markdown_to_excel,
)

# Attempt to import markdown_to_csv; if not found, set to None.
try:
    from xml_process import markdown_to_csv
except ImportError:
    markdown_to_csv = None

class TestXMLProcess(unittest.TestCase):
    def setUp(self):
        # Create a temporary directory for test files.
        self.temp_dir = tempfile.TemporaryDirectory()
        self.data_dir = self.temp_dir.name

    def tearDown(self):
        self.temp_dir.cleanup()

    def test_xml_to_markdown_conversion(self):
        # Create a sample XML content.
        xml_content = '''<?xml version="1.0" encoding="UTF-8"?>
<document>
  <pubinfo>
    <pubtitle>Test Title</pubtitle>
    <edition>1.0</edition>
  </pubinfo>
  <intro>
    <block>
      <para>This is a test paragraph.</para>
    </block>
  </intro>
</document>
'''
        xml_path = os.path.join(self.data_dir, "test.xml")
        with open(xml_path, "w", encoding="utf-8") as f:
            f.write(xml_content)
        
        md_output = xml_to_markdown_conversion(xml_path)
        # Check that the generated Markdown contains the expected text.
        self.assertIn("Test Title", md_output)
        self.assertIn("This is a test paragraph.", md_output)

    
    def test_markdown_to_excel(self):
        # Create a sample Markdown file with a caption and a table.
        md_content = '''*Program Radio for Local Area Frequency*
| Col1 | Col2 |
|------|------|
| A    | B    |
| C    | D    |
'''
        md_path = os.path.join(self.data_dir, "output.md")
        excel_path = os.path.join(self.data_dir, "output.xlsx")
        with open(md_path, "w", encoding="utf-8") as f:
            f.write(md_content)
        
        markdown_to_excel(md_path, excel_path)
        
        # Load the Excel workbook and check its content.
        wb = load_workbook(excel_path)
        ws = wb.active
        
        # The first row should contain the table caption from the italicized line.
        table_caption = ws.cell(row=1, column=1).value
        self.assertEqual(table_caption, "Program Radio for Local Area Frequency")
        
        # Check that the table caption cell is styled correctly.
        # (We expect a black background and white bold text.)
        from openpyxl.styles import PatternFill, Font
        cell = ws.cell(row=1, column=1)
        # openpyxl returns colors as ARGB (e.g., "FF000000" for black)
        self.assertTrue(cell.fill.start_color.rgb.endswith("000000"))
        self.assertTrue(cell.font.color.rgb.endswith("FFFFFF"))
        self.assertTrue(cell.font.bold)
        
        # Next row should be the header row.
        header_row = [ws.cell(row=2, column=i).value for i in range(1, 3)]
        self.assertEqual(header_row, ["Col1", "Col2"])
        
        # And the data rows.
        data_row1 = [ws.cell(row=3, column=i).value for i in range(1, 3)]
        data_row2 = [ws.cell(row=4, column=i).value for i in range(1, 3)]
        self.assertEqual(data_row1, ["A", "B"])
        self.assertEqual(data_row2, ["C", "D"])
        
        # Check that text wrapping is enabled in a sample cell.
        self.assertTrue(ws.cell(row=2, column=1).alignment.wrap_text)

if __name__ == "__main__":
    unittest.main()
